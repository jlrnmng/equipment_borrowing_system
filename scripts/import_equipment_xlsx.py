import argparse
import hashlib
import os
import sys

import openpyxl
import pymysql
import pymysql.cursors

# Ensure project root is importable when running script directly from /scripts.
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

from config.config import config


REQUIRED_HEADERS = {
    'ITEM': 'equipment_name',
    'BRAND': 'brand',
    'SERIAL NUMBER': 'serial_number',
    'INVENTORY NUMBER': 'inventory_number',
    'PROPERTY STOCK': 'property_stock_number',
}

OPTIONAL_HEADERS = {
    'CATEGORY': 'category',
}


def normalize_header(value):
    return str(value or '').strip().upper()


def build_equipment_code(inventory_number):
    """Create a deterministic, compact unique code from inventory number."""
    digest = hashlib.sha1(inventory_number.encode('utf-8')).hexdigest()[:10].upper()
    return f"EQ-{digest}"


def get_connection():
    cfg = config['default']
    return pymysql.connect(
        host=cfg.MYSQL_HOST,
        port=cfg.MYSQL_PORT,
        user=cfg.MYSQL_USER,
        password=cfg.MYSQL_PASSWORD,
        database=cfg.MYSQL_DB,
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=False,
    )


def load_rows(xlsx_path, sheet_name=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    normalized = [normalize_header(h) for h in header_row]
    index_map = {h: i for i, h in enumerate(normalized) if h}

    missing = [h for h in REQUIRED_HEADERS if h not in index_map]
    if missing:
        raise ValueError(f"Missing required headers: {', '.join(missing)}")

    optional_index_map = {
        header: index_map[header]
        for header in OPTIONAL_HEADERS
        if header in index_map
    }

    parsed_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = row[index_map['ITEM']] if index_map['ITEM'] < len(row) else None
        inv_no = row[index_map['INVENTORY NUMBER']] if index_map['INVENTORY NUMBER'] < len(row) else None

        if item is None and inv_no is None:
            continue

        inventory_number = str(inv_no or '').strip()
        category = 'General'
        if 'CATEGORY' in optional_index_map and optional_index_map['CATEGORY'] < len(row):
            category = str(row[optional_index_map['CATEGORY']] or '').strip() or 'General'

        parsed_rows.append(
            {
                'equipment_code': build_equipment_code(inventory_number),
                'equipment_name': str(item or '').strip(),
                'category': category,
                'brand': str(row[index_map['BRAND']] or '').strip() or None,
                'serial_number': str(row[index_map['SERIAL NUMBER']] or '').strip() or None,
                'inventory_number': inventory_number,
                'property_stock_number': str(row[index_map['PROPERTY STOCK']] or '').strip() or None,
            }
        )

    return ws.title, parsed_rows


def upsert_rows(rows):
    sql = """
        INSERT INTO equipment (
            equipment_code,
            equipment_name,
            category,
            brand,
            serial_number,
            inventory_number,
            property_stock_number,
            status,
            condition_status
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, 'available', 'good')
        ON DUPLICATE KEY UPDATE
            equipment_code = VALUES(equipment_code),
            equipment_name = VALUES(equipment_name),
            category = VALUES(category),
            brand = VALUES(brand),
            serial_number = VALUES(serial_number),
            inventory_number = VALUES(inventory_number),
            property_stock_number = VALUES(property_stock_number),
            status = 'available',
            updated_at = CURRENT_TIMESTAMP
    """

    inserted_or_updated = 0
    with get_connection() as conn:
        with conn.cursor() as cursor:
            for row in rows:
                if not row['equipment_name'] or not row['inventory_number']:
                    continue
                cursor.execute(
                    sql,
                    (
                        row['equipment_code'],
                        row['equipment_name'],
                        row['category'],
                        row['brand'],
                        row['serial_number'],
                        row['inventory_number'],
                        row['property_stock_number'],
                    ),
                )
                inserted_or_updated += 1
        conn.commit()

    return inserted_or_updated


def main():
    parser = argparse.ArgumentParser(description='Import equipment inventory from XLSX.')
    parser.add_argument(
        '--file',
        default='asoh_equipment_inventory.xlsx',
        help='Path to xlsx file (default: asoh_equipment_inventory.xlsx)',
    )
    parser.add_argument('--sheet', default=None, help='Optional sheet name')
    args = parser.parse_args()

    if not os.path.exists(args.file):
        raise FileNotFoundError(f'XLSX file not found: {args.file}')

    sheet, rows = load_rows(args.file, args.sheet)
    total = upsert_rows(rows)
    print(f'Sheet: {sheet}')
    print(f'Rows parsed: {len(rows)}')
    print(f'Rows inserted/updated: {total}')


if __name__ == '__main__':
    main()
